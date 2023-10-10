from django.conf import settings
from django.contrib.auth.hashers import make_password
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.generic import ListView, FormView
from customuser.forms import CustomUserQueryForm, CustomUserForm, FileUploadForm
from surveys.models import Ujf, Busyo, Location, Post, CustomUser
from pytz import timezone
import unicodedata
import openpyxl
import pandas as pd

# 部署リスト取得処理 FetchAPI用
# パラメータ：nendo 年度
def getBusyoList(request):
        pnendo = request.POST.get('nendo')
        plistid = request.POST.get('list_id')

        # リスト作成
        busyo_list = llist = [('', '')] + list(Busyo.objects.filter(nendo=pnendo).values_list('bu_code', 'bu_name').order_by('bu_code'))
        bulst = []
        for pdat in busyo_list:
             bulst.append(list(pdat))

        data = {
             'list_id': plistid,
             'rlist': bulst,
        }
        return JsonResponse(data)

# 勤務地リスト取得処理 FetchAPI用
# パラメータ：nendo 年度
def getLocationList(request):
        pnendo = request.POST.get('nendo')
        plistid = request.POST.get('list_id')

        # リスト作成
        location_list = llist = [('', '')] + list(Location.objects.filter(nendo=pnendo).values_list('location_code', 'location_name').order_by('location_code'))
        loclst = []
        for pdat in location_list:
             loclst.append(list(pdat))

        data = {
             'list_id': plistid,
             'rlist': loclst,
        }
        return JsonResponse(data)

# 役職リスト取得処理 FetchAPI用
# パラメータ：nendo 年度
def getPostList(request):
        pnendo = request.POST.get('nendo')
        plistid = request.POST.get('list_id')

        # リスト作成
        post_list = llist = [('', '')] + list(Post.objects.filter(nendo=pnendo).values_list('post_code', 'post_name').order_by('post_code'))
        pstlst = []
        for pdat in post_list:
             pstlst.append(list(pdat))

        data = {
             'list_id': plistid,
             'rlist': pstlst,
        }
        return JsonResponse(data)

# 検索条件(セッション値）から（検索クエリを発行）該当リストを返す
def makeCustomUserList(request):
    form_value = request.session['form_value']
    nendo = form_value[0]
    busyo = form_value[1]
    location = form_value[2]
    post = form_value[3]

    # 検索条件
    exact_nendo = Q()
    exact_busyo = Q()
    exact_location = Q()
    exact_post = Q()
    if len(nendo) != 0 and nendo[0]:
        exact_nendo = Q(nendo__exact=nendo)
    if len(busyo) != 0 and busyo[0]:
        busyoid = Busyo.objects.filter(nendo=nendo, bu_code=busyo)[:1]
        exact_busyo = Q(busyo_id__exact=busyoid)
    if len(location) != 0 and location[0]:
        locationid = Location.objects.filter(nendo=nendo, location_code=location)[:1]
        exact_location = Q(location_id__exact=locationid)
    if len(post) != 0 and post[0]:
        postid = Post.objects.filter(nendo=nendo, post_code=post)[:1]
        exact_post = Q(post_id__exact=postid)

    return (CustomUser.objects.select_related()
            .filter(exact_nendo & exact_busyo & exact_location & exact_post)
            .order_by('busyo_id__bu_code', 'location_id__location_code', 'post_id__post_code', 'user_id')
        )

# ユーザーマスタメンテナンス 一覧
class CUsersListView(LoginRequiredMixin, ListView):
    template_name = 'customuser/customuser_list.html'
    model = CustomUser

    # 検索条件をクリアするためのフラグ
    ini_flg = True

    def post(self, request, *args, **kwargs):
        # 一度検索したので、初期化フラグをOFF
        self.ini_flg = False

        # フォームの内容をセッション情報へ
        form_value = [
            self.request.POST.get('nendo', None),
            self.request.POST.get('busyo', None),
            self.request.POST.get('location', None),
            self.request.POST.get('post', None),
        ]
        request.session['form_value'] = form_value

        # 検索時にページネーションに関連したエラーを防ぐ(ページネーションしないけど一応)
        self.request.GET = self.request.GET.copy()
        self.request.GET.clear()
        return self.get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # 年度リスト作成
        nendo_user = CustomUser.objects.values_list('nendo')
        nendo_ujf = Ujf.objects.filter(key1=1, key2='1').values_list('naiyou4')
        nendo_list = sorted(nendo_user.union(nendo_ujf), key=lambda obj:obj, reverse=True)  # 降順
        nendolst = []
        for i in range(len(nendo_list)):
            nendolst.append((nendo_list[i][0], nendo_list[i][0]))

        nendo = nendo_list[0][0]    # 年度リストの先頭値
        busyo = ''
        location = ''
        post = ''
        # sessionに値がある場合
        if 'form_value' in self.request.session:
            # 初期化フラグがTrueなら
            if self.ini_flg:
                # セッションをクリアする
                del self.request.session['form_value']
            else:
                # 初期化フラグがFalseなら、セッションの値をセットする
                form_value = self.request.session['form_value']
                nendo = form_value[0]
                busyo = form_value[1]
                location = form_value[2]
                post = form_value[3]

        default_data = {'nendo': nendo,
                        'busyo': busyo,
                        'location': location,
                        'post': post,
                        }

        # 部署リスト作成
        blist = [('', '')] + list(Busyo.objects.filter(nendo=nendo).values_list('bu_code', 'bu_name').order_by('bu_code'))

        # 勤務地リスト作成
        llist = [('', '')] + list(Location.objects.filter(nendo=nendo).values_list('location_code', 'location_name').order_by('location_code'))

        # 役職リスト作成
        plist = [('', '')] + list(Post.objects.filter(nendo=nendo).values_list('post_code', 'post_name').order_by('post_code'))

        query_form = CustomUserQueryForm(initial=default_data)  # 検索フォーム

        # リスト
        query_form.fields['nendo'].choices = nendolst
        query_form.fields['busyo'].choices = blist
        query_form.fields['location'].choices = llist
        query_form.fields['post'].choices = plist

        context['query_form'] = query_form
        context['ini_flg'] = self.ini_flg

        return context

    def get_queryset(self):
        # 初期化フラグをURLパラメータから取得
        flg_wrk = self.request.GET.get('ini_flg')
        # パラメータにあったら
        if flg_wrk:
            # 上書き
            if flg_wrk == 'False':
                self.ini_flg = False
            else:
                self.ini_flg = True

        # sessionに値があって、検索ボタン押下なら、セッション値でクエリ発行する。
        if not self.ini_flg and 'form_value' in self.request.session:
            return makeCustomUserList(self.request)
        else:
            # 何も返さない
            return CustomUser.objects.none()

# 新規登録
class CUsersCreateView(LoginRequiredMixin, FormView):
    template_name = 'customuser/customuser_new.html'
    form_class = CustomUserForm
    success_url = '/customusers?ini_flg=False'

    # formにパラメータを渡す為のオーバーライド
    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(CUsersCreateView, self).get_form_kwargs()

        # パラメータ年度、編集モード(新規）をフォームへ渡す
        pnendo = self.kwargs['pnendo']
        mod = "new"
        kwargs.update({'pnendo': pnendo, 'mod': mod})

        return kwargs

    def form_valid(self, form):
        data = form.cleaned_data

        # ユーザマスタに登録
        cuser = CustomUser()
        cuser.nendo = data['nendo']
        cuser.user_id = data['user_id']
        cuser.first_name = data['first_name']
        cuser.last_name = data['last_name']
        cuser.email = data['email']
        cuser.is_staff = data['is_staff']
        cuser.post_id = Post.objects.get(nendo=data['nendo'], post_code=data['post'])
        cuser.busyo_id = Busyo.objects.get(nendo=data['nendo'], bu_code=data['busyo'])
        cuser.location_id = Location.objects.get(nendo=data['nendo'], location_code=data['location'])
        cuser.created_by = self.request.user.username

        cuser.save()

        # ログインマスタに登録があれば更新、なければ登録
        if User.objects.filter(username=cuser.user_id).exists():
            auser = User.objects.get(username=cuser.user_id)
            auser.first_name = cuser.first_name
            auser.last_name = cuser.last_name
            auser.email = cuser.email
            auser.is_staff = cuser.is_staff
            auser.save()
        else:
            new_user = User()
            new_user.username = cuser.user_id
            new_user.password = make_password(cuser.user_id)
            new_user.first_name = cuser.first_name
            new_user.last_name = cuser.last_name
            new_user.email = cuser.email
            new_user.is_staff = cuser.is_staff
            new_user.is_active = True
            new_user.is_superuser = False
            new_user.save()

        return super().form_valid(form)

# 編集・削除
class CUsersEditView(LoginRequiredMixin, FormView):
    template_name = 'customuser/customuser_edit.html'
    form_class = CustomUserForm
    success_url = '/customusers?ini_flg=False'

    # formにパラメータを渡す為のオーバーライド
    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(CUsersEditView, self).get_form_kwargs()

        # パラメータ年度、編集モード(更新）、オブジェクト情報をフォームへ渡す
        pnendo = self.kwargs['pnendo']
        uid = self.kwargs['id']
        mod = "edit"
        uobj = CustomUser.objects.get(id=uid)

        kwargs.update({'pnendo': pnendo, 'mod': mod, 'uobj': uobj})

        return kwargs

    def form_valid(self, form):
        data = form.cleaned_data

        # 更新の時
        if "btn_update" in self.request.POST:
            # ユーザマスタに更新
            cuser = CustomUser.objects.get(nendo=data['nendo'], user_id=data['user_id'])
            cuser.first_name = data['first_name']
            cuser.last_name = data['last_name']
            cuser.email = data['email']
            cuser.is_staff = data['is_staff']
            cuser.post_id = Post.objects.get(nendo=data['nendo'], post_code=data['post'])
            cuser.busyo_id = Busyo.objects.get(nendo=data['nendo'], bu_code=data['busyo'])
            cuser.location_id = Location.objects.get(nendo=data['nendo'], location_code=data['location'])
            cuser.update_by = self.request.user.username

            cuser.save()

            # ログインマスタに登録があれば更新、なければ登録
            if User.objects.filter(username=cuser.user_id).exists():
                auser = User.objects.get(username=cuser.user_id)
                auser.first_name = cuser.first_name
                auser.last_name = cuser.last_name
                auser.email = cuser.email
                auser.is_staff = cuser.is_staff
                auser.save()
            else:
                new_user = User()
                new_user.username = cuser.user_id
                new_user.password = make_password(cuser.user_id)
                new_user.first_name = cuser.first_name
                new_user.last_name = cuser.last_name
                new_user.email = cuser.email
                new_user.is_staff = cuser.is_staff
                new_user.is_active = True
                new_user.is_superuser = False
                new_user.save()
        # 削除の時
        elif "btn_delete" in self.request.POST:
            cuser = CustomUser.objects.get(nendo=data['nendo'], user_id=data['user_id'])
            cuser.delete()

        return super().form_valid(form)

# Excelダウンロード
def download_excel(request):
    wb = openpyxl.load_workbook(str(settings.BASE_DIR) + '/media/customuser.xlsx')
    ws = wb.active

    # データ出力開始行
    row = 2
    # 列数
    col_max = 16

    # スタイルを取得
    cellstylelist = []
    for i in range(col_max):
        cellstyle = ws.cell(row, (i+1))._style
        cellstylelist.append(cellstyle)

    ulist = makeCustomUserList(request)
    for udat in ulist:
        # 値を設定
        ws.cell(row, 1).value = udat.nendo
        ws.cell(row, 2).value = udat.busyo_id.bu_code
        ws.cell(row, 3).value = udat.busyo_id.bu_name
        ws.cell(row, 4).value = udat.location_id.location_code
        ws.cell(row, 5).value = udat.location_id.location_name
        ws.cell(row, 6).value = udat.post_id.post_code
        ws.cell(row, 7).value = udat.post_id.post_name
        ws.cell(row, 8).value = udat.user_id
        ws.cell(row, 9).value = udat.last_name
        ws.cell(row, 10).value = udat.first_name
        ws.cell(row, 11).value = udat.email
        ws.cell(row, 12).value = udat.is_staff
        ws.cell(row, 13).value = udat.created_by
        ws.cell(row, 14).value = udat.created_at.astimezone(timezone('Asia/Tokyo')).strftime('%Y/%m/%d %H:%M:%S')
        ws.cell(row, 15).value = udat.update_by
        ws.cell(row, 16).value = udat.updated_at.astimezone(timezone('Asia/Tokyo')).strftime('%Y/%m/%d %H:%M:%S')
        # スタイルを設定
        for i in range(col_max):
            ws.cell(row, (i+1))._style = cellstylelist[i]
        row += 1
        
    # ダウンロード
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' % 'customuser.xlsx'
    wb.save(response)

    return response

# Excel アップロード
class FileUploadView(LoginRequiredMixin, FormView):
    template_name = 'customuser/fileupload.html'
    form_class = FileUploadForm
    success_url = '/customusers?ini_flg=False'

    def form_valid(self, form):
        # Excel取込み
        df = pd.read_excel(form.cleaned_data['file'], header=0)

        # エラーチェック
        # 組合せチェック用
        budic = dict()
        lodic = dict()
        podic = dict()
        # 行数はIndexが0からで、ヘッダ1行を除くので+2
        for i in range(df.shape[0]) :
            ldat = df.iloc[i]

            # 年度チェック
            strnendo = str(ldat['年度'])
            nendo = int('0')
            if strnendo == 'nan':
                form.add_error(None, '年度の入力がありません。(%s行目)' % str((i+2)))
            elif not strnendo.isdecimal():
                form.add_error(None, '年度に数字以外の文字の入力があります。(%s行目)' % str((i+2)))
            else:
                nendo = int(strnendo)
            
            # 部署チェック
            str_bucode = str(ldat['部署'])
            if str_bucode == 'nan':
                form.add_error(None, '部署コードの入力がありません。(%s行目)' % str((i+2)))
            else:
                # コードが文字列じゃなかったら
                if type(ldat['部署']) is not str:
                    # x.0を消す
                    if str_bucode.endswith('.0'):
                        str_bucode = str_bucode[:(len(str_bucode)-2)]
            str_buname = str(ldat[2])
            if str_buname == 'nan':
                form.add_error(None, '部署名の入力がありません。(%s行目)' % str((i+2)))

            if nendo != 0 and str_bucode != 'nan' and str_buname != 'nan':
                # 年度と部署コードを半角に変換してキーを作成
                bukey = str(nendo) + '_' + unicodedata.normalize('NFKC', str_bucode)
                # キーを検索
                if bukey in budic:
                    # キーが登録済みなのに、名前が違ったらエラー
                    if budic[bukey] != str_buname:
                        form.add_error(None, '異なる部署名の部署コードが存在します。(%s行目)' % str((i+2)))
                else:
                    # 未登録なら登録
                    budic[bukey] = str_buname

            # 勤務地チェック
            str_loccode = str(ldat['勤務地'])
            if str_loccode == 'nan':
                form.add_error(None, '勤務地コードの入力がありません。(%s行目)' % str((i+2)))
            else:
                # コードが文字列じゃなかったら
                if type(ldat['勤務地']) is not str:
                    # x.0を消す
                    if str_loccode.endswith('.0'):
                        str_loccode = str_loccode[:(len(str_loccode)-2)]
            str_locname = str(ldat[4])
            if str_locname == 'nan':
                form.add_error(None, '勤務地名の入力がありません。(%s行目)' % str((i+2)))

            if nendo != 0 and str_loccode != 'nan' and str_locname != 'nan':
                # 年度と勤務地コードを半角に変換してキーを作成
                lockey = str(nendo) + '_' + unicodedata.normalize('NFKC', str_loccode)
                # キーを検索
                if lockey in lodic:
                    # キーが登録済みなのに、名前が違ったらエラー
                    if lodic[lockey] != str_locname:
                        form.add_error(None, '異なる勤務地名の勤務地コードが存在します。(%s行目)' % str((i+2)))
                else:
                    # 未登録なら登録
                    lodic[lockey] = str_locname
            
            # 役職チェック
            str_pstcode = str(ldat['役職'])
            if str_pstcode == 'nan':
                form.add_error(None, '役職コードの入力がありません。(%s行目)' % str((i+2)))
            else:
                # コードが文字列じゃなかったら
                if type(ldat['役職']) is not str:
                    # x.0を消す
                    if str_pstcode.endswith('.0'):
                        str_pstcode = str_pstcode[:(len(str_pstcode)-2)]
            str_pstname = str(ldat[6])
            if str_pstname == 'nan':
                form.add_error(None, '役職名の入力がありません。(%s行目)' % str((i+2)))

            if nendo != 0 and str_pstcode != 'nan' and str_pstname != 'nan':
                # 年度と役職コードを半角に変換してキーを作成
                pstkey = str(nendo) + '_' + unicodedata.normalize('NFKC', str_pstcode)
                # キーを検索
                if pstkey in podic:
                    # キーが登録済みなのに、名前が違ったらエラー
                    if podic[pstkey] != str_pstname:
                        form.add_error(None, '異なる役職名の役職コードが存在します。(%s行目)' % str((i+2)))
                else:
                    # 未登録なら登録
                    podic[pstkey] = str_pstname

            # ユーザ情報必須入力チェック
            str_userid = str(ldat['ユーザーID'])
            if str_userid == 'nan':
                form.add_error(None, 'ユーザーIDの入力がありません。(%s行目)' % str((i+2)))
            str_name1 = str(ldat['氏名'])
            if str_name1 == 'nan':
                form.add_error(None, '氏名の入力がありません。(%s行目)' % str((i+2)))
            str_name2 = str(ldat[9])
            if str_name2 == 'nan':
                form.add_error(None, '氏名の入力がありません。(%s行目)' % str((i+2)))

        # エラーがあったら終了
        if form.non_field_errors():
            return super().form_invalid(form)

        # エラーがなければ登録処理
        # 部署
        for bkey in budic.keys():
            splist = bkey.split('_')
            nendo = splist[0]
            code = splist[1]
            name = budic[bkey]
            # 部署マスタにあれば更新、なければ登録
            if Busyo.objects.filter(nendo=nendo, bu_code=code).exists():
                busyodat = Busyo.objects.get(nendo=nendo, bu_code=code)
                busyodat.bu_name = name
                busyodat.update_by = self.request.user.username
                busyodat.save()
            else:
                busyodat = Busyo()
                busyodat.nendo = nendo
                busyodat.bu_code = code
                busyodat.bu_name = name
                busyodat.created_by = self.request.user.username
                busyodat.save()

        # 勤務地
        for lkey in lodic.keys():
            splist = lkey.split('_')
            nendo = splist[0]
            code = splist[1]
            name = lodic[lkey]
            # 勤務地マスタにあれば更新、なければ登録
            if Location.objects.filter(nendo=nendo, location_code=code).exists():
                locdat = Location.objects.get(nendo=nendo, location_code=code)
                locdat.location_name = name
                locdat.update_by = self.request.user.username
                locdat.save()
            else:
                locdat = Location()
                locdat.nendo = nendo
                locdat.location_code = code
                locdat.location_name = name
                locdat.created_by = self.request.user.username
                locdat.save()

        # 役職
        for pkey in podic.keys():
            splist = pkey.split('_')
            nendo = splist[0]
            code = splist[1]
            name = podic[pkey]
            # 役職マスタにあれば更新、なければ登録
            if Post.objects.filter(nendo=nendo, post_code=code).exists():
                locdat = Post.objects.get(nendo=nendo, post_code=code)
                locdat.post_name = name
                locdat.update_by = self.request.user.username
                locdat.save()
            else:
                locdat = Post()
                locdat.nendo = nendo
                locdat.post_code = code
                locdat.post_name = name
                locdat.created_by = self.request.user.username
                locdat.save()

        # ユーザマスタはファイルから
        for i in range(df.shape[0]) :
            ldat = df.iloc[i]

            nendo = int(ldat['年度'])
            bu_code = int(ldat['部署'])
            loc_code = int(ldat['勤務地'])
            post_code = int(ldat['役職'])
            user_id = ldat['ユーザーID']
            last_name = ldat['氏名']
            first_name = ldat[9]
            email = ldat['メールアドレス']
            if type(email) is not str and str(email) == 'nan':
                email = ''
            staff = False
            if ldat['管理者権限'] == True:
                staff = True

            # ユーザマスタに登録があれば更新、なければ登録
            if CustomUser.objects.filter(nendo=nendo, user_id=user_id).exists():
                cuser = CustomUser.objects.get(nendo=nendo, user_id=user_id)
                cuser.last_name = last_name
                cuser.first_name = first_name
                cuser.email = email
                cuser.is_staff = staff
                cuser.post_id = Post.objects.get(nendo=nendo, post_code=post_code)
                cuser.busyo_id = Busyo.objects.get(nendo=nendo, bu_code=bu_code)
                cuser.location_id = Location.objects.get(nendo=nendo, location_code=loc_code)
                cuser.update_by = self.request.user.username
                cuser.save()
            else:
                cuser = CustomUser()
                cuser.nendo = nendo
                cuser.user_id = user_id
                cuser.last_name = last_name
                cuser.first_name = first_name
                cuser.email = email
                cuser.is_staff = staff
                cuser.post_id = Post.objects.get(nendo=nendo, post_code=post_code)
                cuser.busyo_id = Busyo.objects.get(nendo=nendo, bu_code=bu_code)
                cuser.location_id = Location.objects.get(nendo=nendo, location_code=loc_code)
                cuser.created_by = self.request.user.username
                cuser.save()

            # ログインマスタに登録があれば更新、なければ登録
            if User.objects.filter(username=user_id).exists():
                auser = User.objects.get(username=user_id)
                auser.first_name = first_name
                auser.last_name = last_name
                auser.email = email
                auser.is_staff = staff
                auser.save()
            else:
                new_user = User()
                new_user.username = user_id
                new_user.password = make_password(user_id)
                new_user.first_name = first_name
                new_user.last_name = last_name
                new_user.email = email
                new_user.is_staff = staff
                new_user.is_active = True
                new_user.is_superuser = False
                new_user.save()

        return super().form_valid(form)
