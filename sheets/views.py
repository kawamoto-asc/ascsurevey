from django.conf import settings
from django import forms
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Q, Max
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, FormView
from sheets.consts import INPUT_TYPE_CHOICES, FIELD_TYPE_CHOICES, AGGRE_TYPE_CHOICES, INPUT_URL, AGGRE_URL
from sheets.models import Sheets, Items
from sheets.forms import SheetQueryForm, SheetForm, ItemForm, FileUploadForm
from surveys.models import Ujf, Menu
from pytz import timezone
import openpyxl
import pandas as pd
import re
import unicodedata

FORM_NUM = 1        # フォーム数
FORM_VALUES = {}    # 前回のform_setのPSOT値
SHEET_VALUES = {}   # シート情報

# 検索条件(セッション値）から（検索クエリを発行）該当リストを返す
def makeSheetList(request):
    form_value = request.session['form_value']
    nendo = form_value[0]

    # 検索条件
    exact_nendo = Q()
    if len(nendo) != 0 and nendo[0]:
        exact_nendo = Q(nendo__exact=nendo)

    return (Sheets.objects.select_related()
            .filter(exact_nendo)
            .order_by('dsp_no')
        )

# シートマスタメンテナンス 一覧
class SheetsListView(LoginRequiredMixin, ListView):
    template_name = 'sheets/sheet_list.html'
    model = Sheets

    # 検索条件をクリアするためのフラグ
    ini_flg = True

    def post(self, request, *args, **kwargs):
        # 一度検索したので、初期化フラグをOFF
        self.ini_flg = False

        # フォームの内容をセッション情報へ
        form_value = [
            self.request.POST.get('nendo', None),
        ]
        request.session['form_value'] = form_value

        # 検索時にページネーションに関連したエラーを防ぐ(ページネーションしないけど一応)
        self.request.GET = self.request.GET.copy()
        self.request.GET.clear()
        return self.get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # 年度リスト作成
        nendo_sheet = Sheets.objects.values_list('nendo')
        nendo_ujf = Ujf.objects.filter(key1=1, key2='1').values_list('naiyou4')
        nendo_list = sorted(nendo_sheet.union(nendo_ujf), key=lambda obj:obj, reverse=True)  # 降順
        nendolst = []
        for i in range(len(nendo_list)):
            nendolst.append((nendo_list[i][0], nendo_list[i][0]))

        nendo = nendo_list[0][0]    # 年度リストの先頭値
        # sessionに値がある場合
        if 'form_value' in self.request.session:
            # 初期化フラグがTrueなら
            if self.ini_flg:
                # セッションをクリアする
                del self.request.session['form_value']
            else:
                # 初期化フラグがFalseなら、セッションの値をセットする
                form_value = self.request.session['form_value']
                nendo = form_value[0]

        default_data = {'nendo': nendo,}
        query_form = SheetQueryForm(initial=default_data)  # 検索フォーム

        # リスト
        query_form.fields['nendo'].choices = nendolst

        context['query_form'] = query_form
        context['ini_flg'] = self.ini_flg

        return context

    def get_queryset(self):
        # 初期化フラグをURLパラメータから取得
        flg_wrk = self.request.GET.get('ini_flg')
        # パラメータにあったら
        if flg_wrk:
            # 上書き
            if flg_wrk == 'False':
                self.ini_flg = False
            else:
                self.ini_flg = True

        # sessionに値があって、検索ボタン押下なら、セッション値でクエリ発行する。
        if not self.ini_flg and 'form_value' in self.request.session:
            return makeSheetList(self.request)
        else:
            # 何も返さない
            return Sheets.objects.none()

# 新規登録
class SheetsCreateView(LoginRequiredMixin, FormView):
    template_name = 'sheets/sheet_new.html'
    form_class = SheetForm
    ItemFormSet = forms.formset_factory(
        form = ItemForm,
        extra = 1,
        max_num = 50,
    )
    form_class2 = ItemFormSet

    # 一覧、シート（Session）情報をクリアするフラグ
    ini_flg = True

    # success_urlは動的にパラメータを渡さないといけないのでオーバーライド
    def get_success_url(self):
        url = reverse_lazy(
            'sheet-new',
             kwargs={'pnendo': self.kwargs['pnendo'],
                     'pflg': 'False',
                     }
            )
        return url

    # formにパラメータを渡す為のオーバーライド
    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super().get_form_kwargs()

        # パラメータ年度、編集モード(新規）をフォームへ渡す
        pnendo = self.kwargs['pnendo']
        mod = "new"
        kwargs.update({'pnendo': pnendo, 'mod': mod})

        return kwargs

    # ２個目のフォームを返す為のオーバーライド
    def get_context_data(self, **kwargs):
        global FORM_NUM
        global FORM_VALUES
        global SHEET_VALUES

        # 初期化フラグをパラメータから更新
        if self.kwargs['pflg'] == 'False':
            self.ini_flg = False
        else:
            self.ini_flg = True

        # 初期化フラグからformset,formデータを削除
        if self.ini_flg:
            FORM_NUM = 1
            FORM_VALUES = {}
            SHEET_VALUES = {}

        # ２個目のフォームを渡す
        context = super().get_context_data(**kwargs)

        # フォームデータがあれば
        if FORM_VALUES:
            if self.request.method == 'GET':
                context.update({
                    'form': self.form_class(SHEET_VALUES),
                    'formset': self.form_class2(FORM_VALUES),
                    })
            else:
                context.update({
                    'formset': self.form_class2(FORM_VALUES),
                    })

        # なければ 新規にformset作成
        else:
            context.update({
                'formset': self.form_class2(self.request.GET or None),
                })

        return context

    def post(self, request, *args, **kwargs):
        global FORM_NUM
        global FORM_VALUES
        global SHEET_VALUES

        # 一度Submitしたので、初期化フラグをOFF
        self.ini_flg = False

        # シートデータ項目データ表示用保存
        fsetwork = request.POST.copy()
        fsetdic = {}
        shtdic = {}
        for key, val in fsetwork.items():
            if key.startswith('form-'):
                fsetdic[key] = val
            else:
                shtdic[key] = val
        # 別フォームなので、シート情報と項目情報を分けて保持
        FORM_VALUES = fsetdic
        SHEET_VALUES = shtdic

        # 追加ボタン押下なら
        if 'btn_add' in request.POST:
            FORM_NUM += 1   # formsetのフォーム数インクリメント
            FORM_VALUES['form-TOTAL_FORMS'] = FORM_NUM

        # 削除ボタン押下なら
        if 'btn_del' in request.POST:
            #cnt_frm = 0
            to_cnt = 0
            del_cnt = 0
            newdic = {}
            delkeyist = []
            # FORM_VALUES form Noを削除行をのけて作り直し
            for i in range(FORM_NUM):
                frm_ck_delete = 'form-' + str(i) + '-ck_delete'
                # チェックされていなければ
                if frm_ck_delete not in FORM_VALUES:
                    to_itemno_str = 'form-' + str(to_cnt) + '-item_no'
                    newdic[to_itemno_str] = FORM_VALUES[frm_ck_delete.replace('ck_delete', 'item_no')]
                    newdic[to_itemno_str.replace('item_no', 'content')] = FORM_VALUES[frm_ck_delete.replace('ck_delete', 'content')]
                    newdic[to_itemno_str.replace('item_no', 'input_type')] = FORM_VALUES[frm_ck_delete.replace('ck_delete', 'input_type')]
                    newdic[to_itemno_str.replace('item_no', 'answer')] = FORM_VALUES[frm_ck_delete.replace('ck_delete', 'answer')]
                    newdic[to_itemno_str.replace('item_no', 'haiten')] = FORM_VALUES[frm_ck_delete.replace('ck_delete', 'haiten')]
                    to_cnt += 1
                else :
                    delkeyist.append(frm_ck_delete)
                    del_cnt += 1
            # 項目No.での並べ変えはしない
            for key, val in newdic.items():
                FORM_VALUES[key] = val
            for keystr in delkeyist:
                del FORM_VALUES[keystr]
            FORM_NUM -= del_cnt
            FORM_VALUES['form-TOTAL_FORMS'] = FORM_NUM

        # 登録ボタン押下なら
        if 'btn_entry' in request.POST:
            form = self.form_class(request.POST)
            # 入力チェック
            self.datCheck_new(form, fsetwork)

            # エラーは全部シート部分に出力
            if form.non_field_errors():
                return self.form_invalid(form)
            
            # データ保存
            self.dataEntry(fsetwork)
            return redirect('/sheets?ini_flg=False')

        return super().post(request, args, kwargs)

    # 入力チェック処理
    def datCheck_new(self, form, frmdic):
        # 半角英数字と_@-.だけ入力可のチェック用
        reg = re.compile(r'^[A-Za-z0-9_@\-\.]+$')

        # 数字と-.だけ入力可のチェック用
        rnum = re.compile(r'^[0-9\-\.]+$')

        # シート部分のエラーチェック
        if not frmdic['nendo']:
            form.add_error(None, '年度を入力してください。')

        if not frmdic['sheet_name']:
            form.add_error(None, 'シート名を入力してください。')
        else:
            sheetname = frmdic['sheet_name']
            if reg.match(sheetname) is None:
                form.add_error(None, 'シート名は半角英数字か_-@で入力してください。')
            else:
                nen = frmdic['nendo']
                sheet = Sheets.objects.filter(
                        nendo = nen,
                        sheet_name = sheetname,
                ).exists()
                if sheet:
                    form.add_error(None, nen + 'の' + sheetname + 'は既に登録済みです。')

        if not frmdic['title']:
            form.add_error(None, 'アンケート名を入力してください。')

        # 一問多答形式の場合は項目数は1件のみ
        if frmdic['input_type'] == '2' and FORM_NUM > 1:
            form.add_error(None, '一問多答形式で登録出来る項目数は１件のみです。')

        # 項目リスト部分のエラーチェック
        if FORM_NUM <= 0:
            form.add_error(None, 'アンケート項目がありません。')
        list_ino = []
        for i in range(FORM_NUM):
            if not frmdic['form-' + str(i) + '-item_no']:
                form.add_error(None, '項目Noの入力がありません。(' + str(i + 1) + '行目)')
            else:
                ino = frmdic['form-' + str(i) + '-item_no']
                if ino in list_ino:
                    form.add_error(None, '項目Noが重複しています。(' + str(i + 1) + '行目)')
                else :
                    list_ino.append(ino)

            if not frmdic['form-' + str(i) + '-content']:
                form.add_error(None, '内容の入力がありません。(' + str(i + 1) + '行目)')

            if frmdic['form-' + str(i) + '-haiten']:
                hten = frmdic['form-' + str(i) + '-haiten']
                if rnum.match(hten) is None:
                    form.add_error(None, '配点に数値以外が入力されています。(' + str(i + 1) + '行目)')

    # データ保存処理
    def dataEntry(self, frmdic):
        with transaction.atomic():
            # シート情報を保存
            sheetdat = Sheets()
            sheetdat.nendo = frmdic['nendo']
            sheetdat.sheet_name = frmdic['sheet_name']
            sheetdat.title = frmdic['title']
            # 表示順
            dno = 1
            if frmdic['dsp_no']:
                # 入力があればその数値
                dno = frmdic['dsp_no']
            else:
                # 入力がなくて登録レコードがあればMax＋1
                rno = Sheets.objects.all().count()
                if rno > 0:
                    nodic = Sheets.objects.all().aggregate(Max('dsp_no'))
                    dno = nodic['dsp_no__max'] + 1
            sheetdat.dsp_no = dno
            sheetdat.input_type = frmdic['input_type']
            sheetdat.aggre_type = frmdic['aggre_type']
            sheetdat.remarks1 = frmdic['remarks1']
            sheetdat.remarks2 = frmdic['remarks2']
            rstaff = False
            if 'req_staff' in frmdic:
                rstaff = True
            sheetdat.req_staff = rstaff

            sheetdat.created_by = self.request.user.username
            sheetdat.save()

            # 項目情報保存
            nendo = frmdic['nendo']
            sheetid = Sheets.objects.get(nendo=nendo, sheet_name=frmdic['sheet_name'])
            for i in range(FORM_NUM):
                itemdat = Items()
                itemdat.nendo = nendo
                itemdat.sheet_id = sheetid
                itemdat.item_no = frmdic['form-' + str(i) + '-item_no']
                itemdat.content = frmdic['form-' + str(i) + '-content']
                itemdat.input_type = frmdic['form-' + str(i) + '-input_type']
                itemdat.answer = frmdic['form-' + str(i) + '-answer']
                if frmdic['form-' + str(i) + '-haiten']:
                    itemdat.haiten = float(frmdic['form-' + str(i) + '-haiten'])
                itemdat.created_by = self.request.user.username
                itemdat.save()

            # メニューへの登録
            inpurl = INPUT_URL + frmdic['sheet_name']
            aggurl = AGGRE_URL + frmdic['sheet_name']

            # 入力画面のURLの登録がなければ
            imenu = Menu.objects.filter(
                        url = inpurl,
                ).exists()
            if not imenu:
                imenudat = Menu()
                imenudat.title = frmdic['title']
                imenudat.url = inpurl
                imenudat.kbn = 1
                imenudat.dsp_no = dno
                imenudat.req_staff = False
                imenudat.created_by = self.request.user.username
                imenudat.save()

            # 集計画面のURLの登録がなければ
            amenu = Menu.objects.filter(
                        url = aggurl,
                ).exists()
            if not amenu:
                amenudat = Menu()
                amenudat.title = frmdic['title'] + ' 集計'
                amenudat.url = aggurl
                amenudat.kbn = 2
                amenudat.dsp_no = dno
                amenudat.req_staff = rstaff
                amenudat.created_by = self.request.user.username
                amenudat.save()

# 変更・削除
class SheetsEditView(LoginRequiredMixin, FormView):
    template_name = 'sheets/sheet_edit.html'
    form_class = SheetForm
    ItemFormSet = forms.formset_factory(
        form = ItemForm,
        extra = 0,
        max_num = 50,
    )
    form_class2 = ItemFormSet

    # 一覧、シート（Session）情報をクリアするフラグ
    ini_flg = True

    # success_urlは動的にパラメータを渡さないといけないのでオーバーライド
    def get_success_url(self):
        url = reverse_lazy(
            'sheet-edit',
             kwargs={'pnendo': self.kwargs['pnendo'],
                     'id': self.kwargs['id'],
                     'pflg': 'False',
                     }
            )
        return url

    # formにパラメータを渡す為のオーバーライド
    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super().get_form_kwargs()

        # パラメータ年度、編集モード(新規）をフォームへ渡す
        pnendo = self.kwargs['pnendo']
        mod = "edit"
        sid = self.kwargs['id']
        sobj = Sheets.objects.get(nendo=pnendo, id=sid)

        kwargs.update({'pnendo': pnendo, 'mod': mod, 'sobj': sobj})

        return kwargs

    # ２個目のフォームを返す為のオーバーライド
    def get_context_data(self, **kwargs):
        global FORM_NUM
        global FORM_VALUES
        global SHEET_VALUES

        # 初期化フラグをパラメータから更新
        if self.kwargs['pflg'] == 'False':
            self.ini_flg = False
        else:
            self.ini_flg = True

        # 初期化フラグからformset,formデータを削除
        if self.ini_flg:
            FORM_NUM = 0
            FORM_VALUES = {}
            SHEET_VALUES = {}

        context = super().get_context_data(**kwargs)

        context.update({
            'pnendo': self.kwargs['pnendo'],
            'id': self.kwargs['id'],
            })

        # フォームデータがあれば
        if FORM_VALUES:
            if self.request.method == 'GET':
                context.update({
                    'form': self.form_class(SHEET_VALUES),
                    'formset': self.form_class2(FORM_VALUES),
                    })
            else:
                context.update({
                    'formset': self.form_class2(FORM_VALUES),
                    })

        # なければ 初回表示で、パラメータからformset作成
        else:
            ilst = Items.objects.filter(nendo=self.kwargs['pnendo'], sheet_id=self.kwargs['id']).order_by('item_no')
            FORM_NUM = len(ilst)
            initial_data = []
            for dat in ilst:
                initial_data.append({'item_no': dat.item_no,
                                     'content': dat.content,
                                      'input_type': dat.input_type,
                                      'answer': dat.answer,
                                      'haiten': dat.haiten,
                                    })
            formset = self.form_class2(initial=initial_data)
            context.update({
                'formset': formset,
                })

        return context

    def post(self, request, *args, **kwargs):
        global FORM_NUM
        global FORM_VALUES
        global SHEET_VALUES

        # 一度Submitしたので、初期化フラグをOFF
        self.ini_flg = False

        # シートデータ項目データ表示用保存
        fsetwork = request.POST.copy()
        fsetdic = {}
        shtdic = {}
        for key, val in fsetwork.items():
            if key.startswith('form-'):
                fsetdic[key] = val
            else:
                shtdic[key] = val
        # 別フォームなので、シート情報と項目情報を分けて保持
        FORM_VALUES = fsetdic
        SHEET_VALUES = shtdic

        # 項目追加ボタン押下なら
        if 'btn_add' in request.POST:
            FORM_NUM += 1   # formsetのフォーム数インクリメント
            FORM_VALUES['form-TOTAL_FORMS'] = FORM_NUM

        # 項目削除ボタン押下なら
        if 'btn_del' in request.POST:
            #cnt_frm = 0
            to_cnt = 0
            del_cnt = 0
            newdic = {}
            delkeyist = []
            # FORM_VALUES form Noを削除行をのけて作り直し
            for i in range(FORM_NUM):
                frm_ck_delete = 'form-' + str(i) + '-ck_delete'
                # チェックされていなければ
                if frm_ck_delete not in FORM_VALUES:
                    to_itemno_str = 'form-' + str(to_cnt) + '-item_no'
                    newdic[to_itemno_str] = FORM_VALUES[frm_ck_delete.replace('ck_delete', 'item_no')]
                    newdic[to_itemno_str.replace('item_no', 'content')] = FORM_VALUES[frm_ck_delete.replace('ck_delete', 'content')]
                    newdic[to_itemno_str.replace('item_no', 'input_type')] = FORM_VALUES[frm_ck_delete.replace('ck_delete', 'input_type')]
                    newdic[to_itemno_str.replace('item_no', 'answer')] = FORM_VALUES[frm_ck_delete.replace('ck_delete', 'answer')]
                    newdic[to_itemno_str.replace('item_no', 'haiten')] = FORM_VALUES[frm_ck_delete.replace('ck_delete', 'haiten')]
                    to_cnt += 1
                else :
                    delkeyist.append(frm_ck_delete)
                    del_cnt += 1
            # 項目No.での並べ変えはしない
            for key, val in newdic.items():
                FORM_VALUES[key] = val
            for keystr in delkeyist:
                del FORM_VALUES[keystr]
            FORM_NUM -= del_cnt
            FORM_VALUES['form-TOTAL_FORMS'] = FORM_NUM

        # 更新ボタン押下なら
        if 'btn_update' in request.POST:
            form = self.form_class(request.POST)
            # 入力チェック
            self.datCheck_update(form, fsetwork)

            # エラーは全部シート部分に出力
            if form.non_field_errors():
                return self.form_invalid(form)
            
            # データ保存
            self.dataEntry(fsetwork)
            return redirect('/sheets?ini_flg=False')
        
        # 削除ボタン押下なら
        if 'btn_delete' in request.POST:
            self.dataDelete()
            return redirect('/sheets?ini_flg=False')

        return super().post(request, args, kwargs)

    # 入力チェック処理
    def datCheck_update(self, form, frmdic):
        # 数字と-.だけ入力可のチェック用
        rnum = re.compile(r'^[0-9\-\.]+$')

        if not frmdic['title']:
            form.add_error(None, 'アンケート名を入力してください。')

        # 一問多答形式の場合は項目数は1件のみ
        if frmdic['input_type'] == '2' and FORM_NUM > 1:
            form.add_error(None, '一問多答形式で登録出来る項目数は１件のみです。')

        # 項目リスト部分のエラーチェック
        if FORM_NUM <= 0:
            form.add_error(None, 'アンケート項目がありません。')
        list_ino = []
        for i in range(FORM_NUM):
            if not frmdic['form-' + str(i) + '-item_no']:
                form.add_error(None, '項目Noの入力がありません(' + str(i + 1) + '行目)')
            else:
                ino = frmdic['form-' + str(i) + '-item_no']
                if ino in list_ino:
                    form.add_error(None, '項目Noが重複しています(' + str(i + 1) + '行目)')
                else :
                    list_ino.append(ino)

            if not frmdic['form-' + str(i) + '-content']:
                form.add_error(None, '内容の入力がありません(' + str(i + 1) + '行目)')

            if frmdic['form-' + str(i) + '-haiten']:
                hten = frmdic['form-' + str(i) + '-haiten']
                if rnum.match(hten) is None:
                    form.add_error(None, '配点に数値以外が入力されています。(' + str(i + 1) + '行目)')

    # データ保存処理
    def dataEntry(self, frmdic):
        with transaction.atomic():
            nendo = self.kwargs['pnendo']
            sheet_id=self.kwargs['id']

            # シート情報を保存
            sheetdat = Sheets.objects.get(nendo=nendo, id=sheet_id)
            sheetdat.title = frmdic['title']
            # 表示順
            dno = 1
            if frmdic['dsp_no']:
                # 入力があればその数値
                dno = frmdic['dsp_no']
            else:
                # 入力がなくて登録レコードがあればMax＋1
                rno = Sheets.objects.all().count()
                if rno > 0:
                    dno = Sheets.objects.all().aggregate(Max('dsp_no')) + 1
            sheetdat.dsp_no = dno
            sheetdat.input_type = frmdic['input_type']
            sheetdat.aggre_type = frmdic['aggre_type']
            sheetdat.remarks1 = frmdic['remarks1']
            sheetdat.remarks2 = frmdic['remarks2']
            rstaff = False
            if 'req_staff' in frmdic:
                rstaff = True
            sheetdat.req_staff = rstaff

            sheetdat.update_by = self.request.user.username
            sheetdat.save()

            # 項目情報保存 delete&Insert
            items = Items.objects.filter(nendo=nendo, sheet_id=sheet_id)
            items.delete()
            for i in range(FORM_NUM):
                itemdat = Items()
                itemdat.nendo = nendo
                itemdat.sheet_id = sheetdat
                itemdat.item_no = frmdic['form-' + str(i) + '-item_no']
                itemdat.content = frmdic['form-' + str(i) + '-content']
                itemdat.input_type = frmdic['form-' + str(i) + '-input_type']
                itemdat.answer = frmdic['form-' + str(i) + '-answer']
                if frmdic['form-' + str(i) + '-haiten']:
                    itemdat.haiten = float(frmdic['form-' + str(i) + '-haiten'])
                itemdat.created_by = self.request.user.username
                itemdat.save()

            # メニューへの更新
            inpurl = INPUT_URL + frmdic['sheet_name']
            aggurl = AGGRE_URL + frmdic['sheet_name']

            # 入力画面のURLの登録があれば
            imenu = Menu.objects.filter(
                        url = inpurl,
                ).exists()
            if imenu:
                imenudat = Menu.objects.get(url=inpurl)
                imenudat.title = frmdic['title']
                imenudat.kbn = 1
                imenudat.dsp_no = dno
                imenudat.req_staff = False
                imenudat.update_by = self.request.user.username
                imenudat.save()

            # 集計画面のURLの登録があれば
            amenu = Menu.objects.filter(
                        url = aggurl,
                ).exists()
            if amenu:
                amenudat = Menu.objects.get(url=aggurl)
                amenudat.title = frmdic['title'] + ' 集計'
                amenudat.kbn = 2
                amenudat.dsp_no = dno
                amenudat.req_staff = rstaff
                amenudat.update_by = self.request.user.username
                amenudat.save()

    # データ削除処理
    def dataDelete(self):
        with transaction.atomic():
            nendo = self.kwargs['pnendo']
            sheet_id=self.kwargs['id']

            # シート情報削除
            sheet = Sheets.objects.get(nendo=nendo, id=sheet_id)
            sheet.delete()

            # 項目情報削除
            items = Items.objects.filter(nendo=nendo, sheet_id=sheet_id)
            items.delete()

# Excelダウンロード
def SheetsDownloadExcel(request, pnendo, id):
    wb = openpyxl.load_workbook(str(settings.BASE_DIR) + '/media/sheet.xlsx')
    ws = wb.active

    # データ出力開始行
    row = 2
    # 列数
    col_max = 21

    # スタイルを取得
    cellstylelist = []
    for i in range(col_max):
        cellstyle = ws.cell(row, (i+1))._style
        cellstylelist.append(cellstyle)

    sheet = Sheets.objects.get(id=id)
    itemlist = Items.objects.filter(nendo=pnendo, sheet_id=sheet)

    # セレクトボックス値からのdicを作成
    ityp_dic = {}
    for ival, ilabel in INPUT_TYPE_CHOICES:
        ityp_dic[ival] = ilabel
    ftyp_dic = {}
    for fval, flabel in FIELD_TYPE_CHOICES:
        ftyp_dic[fval] = flabel
    atyp_dic = {}
    for aval, alabel in AGGRE_TYPE_CHOICES:
        atyp_dic[aval] = alabel

    for idat in itemlist:
        # 値を設定
        ws.cell(row, 1).value = sheet.nendo
        ws.cell(row, 2).value = sheet.sheet_name
        ws.cell(row, 3).value = sheet.title
        ws.cell(row, 4).value = sheet.dsp_no
        ws.cell(row, 5).value = sheet.input_type
        ws.cell(row, 6).value = ityp_dic[sheet.input_type]
        ws.cell(row, 7).value = sheet.aggre_type
        ws.cell(row, 8).value = atyp_dic[sheet.aggre_type]
        ws.cell(row, 9).value = sheet.req_staff
        ws.cell(row, 10).value = sheet.remarks1
        ws.cell(row, 11).value = sheet.remarks2
        ws.cell(row, 12).value = idat.item_no
        ws.cell(row, 13).value = idat.content
        ws.cell(row, 14).value = idat.input_type
        ws.cell(row, 15).value = ftyp_dic[idat.input_type]
        ws.cell(row, 16).value = idat.answer
        ws.cell(row, 17).value = idat.haiten
        # 作成・更新情報はシートのものを出力（項目はDeleteInsertだから作成情報＝更新情報になってしまうから）
        ws.cell(row, 18).value = sheet.created_by
        ws.cell(row, 19).value = sheet.created_at.astimezone(timezone('Asia/Tokyo')).strftime('%Y/%m/%d %H:%M:%S')
        ws.cell(row, 20).value = sheet.update_by
        ws.cell(row, 21).value = sheet.updated_at.astimezone(timezone('Asia/Tokyo')).strftime('%Y/%m/%d %H:%M:%S')
        # スタイルを設定
        for i in range(col_max):
            ws.cell(row, (i+1))._style = cellstylelist[i]
        row += 1
        
    # ダウンロード
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' % 'sheet.xlsx'
    wb.save(response)

    return response

# Excel アップロード
class SheetFileUploadView(LoginRequiredMixin, FormView):
    template_name = 'sheets/fileupload.html'
    form_class = FileUploadForm
    success_url = '/sheets?ini_flg=False'

    def form_valid(self, form):
        # Excel取込み
        df = pd.read_excel(form.cleaned_data['file'], header=0)

        # 数字と-.だけ入力可のチェック用
        reg = re.compile(r'^[0-9\-\.]+$')

        # セレクトボックス値からのdicを作成
        ityp_dic = {}
        for ival, ilabel in INPUT_TYPE_CHOICES:
            ityp_dic[ival] = ilabel
        ftyp_dic = {}
        for fval, flabel in FIELD_TYPE_CHOICES:
            ftyp_dic[fval] = flabel
        atyp_dic = {}
        for aval, alabel in AGGRE_TYPE_CHOICES:
            atyp_dic[aval] = alabel

        # エラーチェック
        chk_nendo = None
        chk_sname = None
        chk_aname = None
        chk_itype = None
        chk_atype = None
        chk_dspno = None
        chk_staff = None
        chk_rmks1 = None
        chk_rmks2 = None
        # 登録済みチェック用
        inolist = []

        # 行数はIndexが0からで、ヘッダ1行を除くので+2
        for i in range(df.shape[0]) :
            ldat = df.iloc[i]

            # 年度チェック
            strnendo = str(ldat['年度'])
            if strnendo == 'nan':
                form.add_error(None, '年度の入力がありません。(%s行目)' % str((i+2)))
            elif reg.match(strnendo) is None:
                form.add_error(None, '年度に数字以外の文字の入力があります。(%s行目)' % str((i+2)))
            else:
                # 比較用年度に値がなければセット
                if not chk_nendo:
                    chk_nendo = int(strnendo)
                else:
                    # 比較用年度と違う値が入っていればエラー
                    if chk_nendo != int(strnendo):
                        form.add_error(None, '異なる年度の値が入力されています。(%s行目)' % str((i+2)))
            
            # シート名チェック
            str_sname = str(ldat['シート名'])
            if str_sname == 'nan':
                form.add_error(None, 'シート名の入力がありません。(%s行目)' % str((i+2)))
            else:
                # 比較用シート名に値がなければセット
                if not chk_sname:
                    chk_sname = str_sname
                else:
                    # 比較用シート名と違う値が入っていればエラー
                    if chk_sname != str_sname:
                        form.add_error(None, '異なるシート名が入力されています。(%s行目)' % str((i+2)))

            # アンケート名チェック
            str_aname = str(ldat['アンケート名'])
            if str_aname == 'nan':
                form.add_error(None, 'アンケート名の入力がありません。(%s行目)' % str((i+2)))
            else:
                # 比較用アンケート名に値がなければセット
                if not chk_aname:
                    chk_aname = str_aname
                else:
                    # 比較用アンケート名と違う値が入っていればエラー
                    if chk_aname != str_aname:
                        form.add_error(None, '異なるアンケート名が入力されています。(%s行目)' % str((i+2)))

            # 表示順チェック
            strdspno = str(ldat['表示順'])
            # 表示順は入力なくてもOK
            if strdspno != 'nan':
                # 入力があるけど数字じゃなければ
                if reg.match(strdspno) is None:
                    form.add_error(None, '表示順に数字以外の文字の入力があります。(%s行目)' % str((i+2)))
                else:
                    # 比較用表示順に値がなければセット
                    if not chk_dspno:
                        chk_dspno = int(float(strdspno))
                    else:
                        # 比較用表示順と違う値が入っていればエラー
                        if chk_dspno != int(float(strdspno)):
                            form.add_error(None, '異なる表示順の値が入力されています。(%s行目)' % str((i+2)))

            # 入力形式チェック
            str_itype = str(ldat['入力形式'])
            if str_itype == 'nan':
                form.add_error(None, '入力形式の入力がありません。(%s行目)' % str((i+2)))
            elif reg.match(str_itype) is None:
                form.add_error(None, '入力形式に誤りがあります。(%s行目)' % str((i+2)))
            else:
                # 入力形式チェック 桁の小さい数値は小数点になって取り込まれるのでfloatに変換後intへ変換
                if int(float(str_itype)) not in ityp_dic:
                    form.add_error(None, '入力形式に誤りがあります。(%s行目)' % str((i+2)))
                else:
                    # 比較用入力形式に値がなければセット
                    if not chk_itype:
                        chk_itype = str_itype
                    else:
                        # 比較用入力形式と違う値が入っていればエラー
                        if chk_itype != str_itype:
                            form.add_error(None, '異なる入力形式が入力されています。(%s行目)' % str((i+2)))
                        
            # 集計タイプチェック
            str_atype = str(ldat['集計タイプ'])
            if str_atype == 'nan':
                form.add_error(None, '集計タイプの入力がありません。(%s行目)' % str((i+2)))
            elif reg.match(str_atype) is None:
                form.add_error(None, '集計タイプに誤りがあります。(%s行目)' % str((i+2)))
            else:
                # 集計タイプチェック 桁の小さい数値は小数点になって取り込まれるのでfloatに変換後intへ変換
                if int(float(str_atype)) not in ityp_dic:
                    form.add_error(None, '集計タイプに誤りがあります。(%s行目)' % str((i+2)))
                else:
                    # 比較用集計タイプに値がなければセット
                    if not chk_atype:
                        chk_atype = str_atype
                    else:
                        # 比較用集計タイプと違う値が入っていればエラー
                        if chk_atype != str_atype:
                            form.add_error(None, '異なる集計タイプが入力されています。(%s行目)' % str((i+2)))
                        
            # 集計画面管理者権限要チェック
            strstaff = str(ldat['集計画面管理者権限要'])
            # 集計画面管理者権限要は入力なくてもOK
            if strstaff != 'nan':
                # 比較用集計画面管理者権限要に値がなければセット
                if not chk_staff:
                    chk_staff = strstaff
                else:
                    # 比較用集計画面管理者権限要と違う値が入っていればエラー
                    if chk_staff != strstaff:
                        form.add_error(None, '異なる集計画面管理者権限要の値が入力されています。(%s行目)' % str((i+2)))

            # 備考1 備考はノーチェックで一番最後に入った値を登録するようにする
            str_rmks1 = str(ldat['備考1'])
            if str_rmks1 != 'nan':
                # 比較用集計タイプに値がなければセット
                if not chk_rmks1:
                    chk_rmks1 = str_rmks1

            # 備考2
            str_rmks2 = str(ldat['備考2'])
            if str_rmks2 != 'nan':
                # 比較用集計タイプに値がなければセット
                if not chk_rmks2:
                    chk_rmks2 = str_rmks2

            # 項目No.入力チェック
            str_itemno = str(ldat['項目No.'])
            if str_itemno == 'nan':
                form.add_error(None, '項目No.の入力がありません。(%s行目)' % str((i+2)))
            else:
                # 入力があるけど数字じゃなければ
                if reg.match(str_itemno) is None:
                    form.add_error(None, '項目No.に数字以外の文字の入力があります。(%s行目)' % str((i+2)))
                # 登録済みチェック
                elif str_itemno in inolist:
                    form.add_error(None, '項目No.が重複しています。(%s行目)' % str((i+2)))
                else:
                    # 登録済みでなければチェック用リストに追加
                    inolist.append(str_itemno)

            # 内容チェック
            str_content = str(ldat['内容'])
            if str_content == 'nan':
                form.add_error(None, '内容の入力がありません。(%s行目)' % str((i+2)))

            # 入力タイプチェック
            str_ftype = str(ldat['入力タイプ'])
            if str_ftype == 'nan':
                form.add_error(None, '入力タイプの入力がありません。(%s行目)' % str((i+2)))
            else:
                # 入力タイプチェック
                if str_ftype not in ftyp_dic:
                    form.add_error(None, '入力タイプに誤りがあります。(%s行目)' % str((i+2)))
            # 解答はノーチェック

            # 配点チェック
            strhten  = str(ldat['配点'])
            if strhten != 'nan':
               if reg.match(strhten) is None:
                form.add_error(None, '配点に数字以外の文字の入力があります。(%s行目)' % str((i+2)))

        # エラーがあったら終了
        if form.non_field_errors():
            return super().form_invalid(form)

        # エラーがなければ登録処理
        reg_staff = False
        if chk_staff == 'True':
            reg_staff = True

        # 表示順の入力がなければ
        if chk_dspno == None:
            # 登録レコードがあればMax＋1
            rno = Sheets.objects.all().count()
            if rno > 0:
                nodic = Sheets.objects.all().aggregate(Max('dsp_no'))
                chk_dspno = nodic['dsp_no__max'] + 1
            else:
                chk_dspno = 1

        with transaction.atomic():
            # シートマスタにあれば更新、なければ登録
            if Sheets.objects.filter(nendo=chk_nendo, sheet_name=chk_sname).exists():
                sheetdat = Sheets.objects.get(nendo=chk_nendo, sheet_name=chk_sname)
                sheetdat.update_by = self.request.user.username
            else:
                sheetdat = Sheets()
                sheetdat.nendo = chk_nendo
                sheetdat.created_by = self.request.user.username

            sheetdat.sheet_name = chk_sname
            sheetdat.title = chk_aname
            sheetdat.dsp_no = chk_dspno
            sheetdat.input_type = chk_itype
            sheetdat.aggre_type = chk_atype
            sheetdat.req_staff = reg_staff
            sheetdat.remarks1 = chk_rmks1
            sheetdat.remarks2 = chk_rmks2
            sheetdat.save()

            # メニューマスタに登録があれば更新、なければ登録
            inpurl = INPUT_URL + chk_sname
            aggurl = AGGRE_URL + chk_sname

            # 入力画面のURL
            imenu = Menu.objects.filter(
                        url = inpurl,
                ).exists()
            if imenu:
                imenudat = Menu.objects.get(url=inpurl)
                imenudat.update_by = self.request.user.username
            else:
                imenudat = Menu()
                imenudat.url = inpurl
                imenudat.created_by = self.request.user.username
            imenudat.title = chk_aname
            imenudat.kbn = 1
            imenudat.dsp_no = chk_dspno
            imenudat.req_staff = False
            imenudat.save()

            # 集計画面のURL
            amenu = Menu.objects.filter(
                        url = aggurl,
                ).exists()
            if amenu:
                amenudat = Menu.objects.get(url=aggurl)
                amenudat.update_by = self.request.user.username
            else:
                amenudat = Menu()
                amenudat.url = aggurl
                amenudat.created_by = self.request.user.username
            amenudat.title = chk_aname + ' 集計'
            amenudat.kbn = 2
            amenudat.dsp_no = chk_dspno
            amenudat.req_staff = reg_staff
            amenudat.save()

            # 項目マスタはファイルからdelete&Insert
            sheet = Sheets.objects.get(nendo=chk_nendo, sheet_name=chk_sname)
            if Items.objects.filter(nendo=chk_nendo, sheet_id=sheet).exists():
                Items.objects.filter(nendo=chk_nendo, sheet_id=sheet).delete()

            for i in range(df.shape[0]) :
                ldat = df.iloc[i]

                item_no = int(float(ldat['項目No.']))
                content = ldat['内容']
                input_type = ldat['入力タイプ']
                answer = ldat['解答']
                haiten = str(ldat['配点'])

                item = Items()
                item.nendo = chk_nendo
                item.sheet_id = sheet
                item.item_no = item_no
                item.content = content
                item.input_type = input_type
                item.answer = answer
                if haiten != 'nan':
                    item.haiten = float(haiten)
                item.created_by = self.request.user.username
                item.save()

        return super().form_valid(form)
