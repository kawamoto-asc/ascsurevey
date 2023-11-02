from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.http import HttpResponse
from django.views.generic import ListView

from sheets.consts import MSUM_STR
from sheets.models import Sheets
from summarize.forms import SummarizeQueryForm
from surveys.models import Ujf, Busyo, Location, Post, Score

from pytz import timezone
import openpyxl

# 検索条件(セッション値）から（検索クエリを発行）該当リストを返す
def makeSummarizeList(request):
    form_value = request.session['form_value']
    nendo = form_value[0]
    busyo = form_value[1]
    location = form_value[2]
    post = form_value[3]
    shname = form_value[4]

    # 検索条件
    exact_nendo = Q()
    exact_busyo = Q()
    exact_location = Q()
    exact_post = Q()
    exact_sheet = Q(sheet_id__sheet_name__exact=shname)
    if len(nendo) != 0 and nendo[0]:
        exact_nendo = Q(nendo__exact=nendo)
    if len(busyo) != 0 and busyo[0]:
        busyoid = Busyo.objects.filter(nendo=nendo, bu_code=busyo)[:1]
        exact_busyo = Q(user_id__busyo_id__exact=busyoid)
    if len(location) != 0 and location[0]:
        locationid = Location.objects.filter(nendo=nendo, location_code=location)[:1]
        exact_location = Q(user_id__location_id__exact=locationid)
    if len(post) != 0 and post[0]:
        postid = Post.objects.filter(nendo=nendo, post_code=post)[:1]
        exact_post = Q(user_id__post_id__exact=postid)


    return (Score.objects.select_related()
            .filter(exact_nendo & exact_busyo & exact_location & exact_post & exact_sheet)
            .order_by('user_id__busyo_id__bu_code', 'user_id__location_id__location_code', 'user_id__post_id__post_code', 'user_id', 'item_id', 'dsp_no')
        )

# 集約一覧
class SummarizeListView(LoginRequiredMixin, ListView):
    template_name = 'summarize/summarize_list.html'
    model = Score

    # 検索条件をクリアするためのフラグ
    ini_flg = True

    def post(self, request, *args, **kwargs):
        # 一度検索したので、初期化フラグをOFF
        self.ini_flg = False

        # フォームの内容をセッション情報へ
        form_value = [
            self.request.POST.get('nendo', None),
            self.request.POST.get('busyo', None),
            self.request.POST.get('location', None),
            self.request.POST.get('post', None),
            self.request.POST.get('shname', None),
        ]
        request.session['form_value'] = form_value

        # 検索時にページネーションに関連したエラーを防ぐ(ページネーションしないけど一応)
        self.request.GET = self.request.GET.copy()
        self.request.GET.clear()
        return self.get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # シート情報を渡す
        qsheet = Sheets.objects.filter(sheet_name=self.kwargs['pshname'],).order_by('-nendo')
        context['qsheet'] = qsheet[0]
        context['subtitle'] = MSUM_STR

        # 年度リスト作成
        nendo_score = Score.objects.values_list('nendo')
        nendo_ujf = Ujf.objects.filter(key1=1, key2='1').values_list('naiyou4')
        nendo_list = sorted(nendo_score.union(nendo_ujf), key=lambda obj:obj, reverse=True)  # 降順
        nendolst = []
        for i in range(len(nendo_list)):
            nendolst.append((nendo_list[i][0], nendo_list[i][0]))

        nendo = nendo_list[0][0]    # 年度リストの先頭値
        busyo = ''
        location = ''
        post = ''
        shname = self.kwargs['pshname']
        # sessionに値がある場合
        if 'form_value' in self.request.session:
            # 初期化フラグがTrueなら
            if self.ini_flg:
                # セッションをクリアする
                del self.request.session['form_value']
            else:
                # 初期化フラグがFalseなら、セッションの値をセットする
                form_value = self.request.session['form_value']
                nendo = form_value[0]
                busyo = form_value[1]
                location = form_value[2]
                post = form_value[3]
                shname = form_value[4]

        default_data = {'nendo': nendo,
                        'busyo': busyo,
                        'location': location,
                        'post': post,
                        'shname': shname,
                        }

        # 部署リスト作成
        blist = [('', '')] + list(Busyo.objects.filter(nendo=nendo).values_list('bu_code', 'bu_name').order_by('bu_code'))

        # 勤務地リスト作成
        llist = [('', '')] + list(Location.objects.filter(nendo=nendo).values_list('location_code', 'location_name').order_by('location_code'))

        # 役職リスト作成
        plist = [('', '')] + list(Post.objects.filter(nendo=nendo).values_list('post_code', 'post_name').order_by('post_code'))

        query_form = SummarizeQueryForm(initial=default_data)  # 検索フォーム

        # リスト
        query_form.fields['nendo'].choices = nendolst
        query_form.fields['busyo'].choices = blist
        query_form.fields['location'].choices = llist
        query_form.fields['post'].choices = plist
        query_form.fields['shname'].initial = shname

        context['query_form'] = query_form
        context['ini_flg'] = self.ini_flg

        return context

    def get_queryset(self):
        # 初期化フラグをURLパラメータから取得
        flg_wrk = self.request.GET.get('ini_flg')
        # パラメータにあったら
        if flg_wrk:
            # 上書き
            if flg_wrk == 'False':
                self.ini_flg = False
            else:
                self.ini_flg = True

        # sessionに値があって、検索ボタン押下なら、セッション値でクエリ発行する。
        if not self.ini_flg and 'form_value' in self.request.session:
            return makeSummarizeList(self.request)
        else:
            # 何も返さない
            return Score.objects.none()

# Excelダウンロード
def dl_score_excel2(request):
    wb = openpyxl.load_workbook(str(settings.BASE_DIR) + '/media/summarize2.xlsx')
    ws = wb.active

    # データ出力開始行
    row = 4
    # 列数
    col_max = 7

    # スタイルを取得
    cellstylelist = []
    for i in range(col_max):
        cellstyle = ws.cell(row, (i+1))._style
        cellstylelist.append(cellstyle)

    sclist = makeSummarizeList(request)

    # タイトルを出力
    ws.cell(1, 1).value = sclist[0].sheet_id.title

    for scdat in sclist:
        # 値を設定
        ws.cell(row, 1).value = scdat.nendo
        ws.cell(row, 2).value = scdat.user_id.busyo_id.bu_name
        ws.cell(row, 3).value = scdat.user_id.location_id.location_name
        ws.cell(row, 4).value = scdat.user_id.post_id.post_name
        ws.cell(row, 5).value = scdat.user_id.last_name + ' ' + scdat.user_id.first_name
        ws.cell(row, 6).value = scdat.inp_data
        ws.cell(row, 7).value = scdat.updated_at.astimezone(timezone('Asia/Tokyo')).strftime('%Y/%m/%d %H:%M:%S')
        # スタイルを設定
        for i in range(col_max):
            ws.cell(row, (i+1))._style = cellstylelist[i]
        row += 1
        
    # ダウンロード
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' % 'summarize.xlsx'
    wb.save(response)

    return response